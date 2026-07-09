import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def export_to_excel(data: list[dict], output_path: str, selected_columns: list[str]) -> bool:
    if not data or not selected_columns:
        return False
        
    try:
        # Create default export via pandas
        df = pd.DataFrame(data)
        df_filtered = df.reindex(columns=selected_columns).fillna("null")
        df_filtered = df_filtered.replace("", "null")
        df_filtered.to_excel(output_path, index=False)
        
        # Load exported file to apply advanced styling with openpyxl
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Define clean, professional color fills
        # We alternate column colors: white and a very soft light blue-gray
        fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        fill_tint = PatternFill(start_color="F2F7FA", end_color="F2F7FA", fill_type="solid")
        
        # Elegant Slate Blue header
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        
        # Fonts
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell_font = Font(name="Segoe UI", size=10)
        
        # Light grey borders
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        # Style Headers
        ws.row_dimensions[1].height = 28
        for col_idx in range(1, len(selected_columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            
        # Style Data Cells (with alternating column backgrounds)
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 20
            for col_idx in range(1, len(selected_columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = cell_font
                cell.border = thin_border
                
                # Apply column-based alternating colors
                if col_idx % 2 == 0:
                    cell.fill = fill_tint
                else:
                    cell.fill = fill_white
                
                # Setup specific alignments based on column types
                col_name = selected_columns[col_idx - 1]
                # Center numeric, short codes, and dates; left-align text
                if col_name in ["STT", "No.", "Q'ty", "Booking No", "Số Booking", "Block", "Equipment Type", "Loại cont", "ETD", "Ngày tàu chạy"]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
        # Auto-adjust column widths based on maximum contents length
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                # Handle potential line breaks in headers
                lines = val.split('\n')
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        wb.save(output_path)
        return True
    except Exception as e:
        print(f"Error styling exported Excel: {e}")
        return False
